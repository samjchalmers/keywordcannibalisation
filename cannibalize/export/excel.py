from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from cannibalize.db.store import CannibalizationCase, Store

HEADERS = [
    "id",
    "query",
    "urls",
    "case_type",
    "severity",
    "similarity",
    "volatility",
    "click_loss",
    "keep_url",
    "recommendation",
    "status",
]


def _row(case: CannibalizationCase) -> list:
    return [
        case.id,
        case.query,
        " | ".join(case.urls),
        case.case_type,
        case.severity_score,
        case.similarity_score,
        case.position_volatility,
        case.estimated_click_loss,
        case.keep_url,
        case.recommendation,
        case.status,
    ]


def _write_sheet(ws, cases: list[CannibalizationCase]) -> None:
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for c in cases:
        ws.append(_row(c))

    if cases:
        end = len(cases) + 1
        ws.conditional_formatting.add(
            f"E2:E{end}",
            ColorScaleRule(
                start_type="min",
                start_color="63BE7B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="F8696B",
            ),
        )

    for col_idx, header in enumerate(HEADERS, start=1):
        max_len = len(header)
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


def export_cases_excel(
    cases: list[CannibalizationCase], path: Path, store: Store | None = None
) -> None:
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    type_counts = Counter(c.case_type or "UNKNOWN" for c in cases)
    total_loss = sum(c.estimated_click_loss or 0.0 for c in cases)
    summary.append(["Metric", "Value"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    summary.append(["Total cases", len(cases)])
    summary.append(["Total estimated click loss", round(total_loss, 1)])
    summary.append([])
    summary.append(["Case type", "Count"])
    for t, n in type_counts.most_common():
        summary.append([t, n])
    summary.append([])
    summary.append(["Top 10 by severity"])
    top = sorted(
        cases, key=lambda c: c.severity_score or 0.0, reverse=True
    )[:10]
    summary.append(["query", "case_type", "severity", "click_loss"])
    for c in top:
        summary.append(
            [c.query, c.case_type, c.severity_score, c.estimated_click_loss]
        )
    summary.column_dimensions["A"].width = 40
    summary.column_dimensions["B"].width = 20

    ws_all = wb.create_sheet("All Cases")
    _write_sheet(ws_all, cases)

    by_type: dict[str, list[CannibalizationCase]] = {}
    for c in cases:
        by_type.setdefault(c.case_type or "UNKNOWN", []).append(c)
    for t, group in by_type.items():
        ws = wb.create_sheet(t[:31])
        _write_sheet(ws, group)

    fixed_cases = [c for c in cases if c.status == "fixed"]
    if fixed_cases:
        ws_fix = wb.create_sheet("Fix Results")
        ws_fix.append(
            ["id", "query", "case_type", "fixed_at", "recommendation"]
        )
        for cell in ws_fix[1]:
            cell.font = Font(bold=True)
        for c in fixed_cases:
            ws_fix.append(
                [c.id, c.query, c.case_type, c.fixed_at, c.recommendation]
            )

    wb.save(path)
